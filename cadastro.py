from tkinter import messagebox
import tkinter as tk 
import openpyxl

def salvar_dados():

    workbook = openpyxl.load_workbook('Vazio.xlsx')
    vendas = workbook['Produtos']

    # Salvar os dados
    cliente = entrada_cliente.get()
    produto = entrada_produto.get()
    quantidade = entrada_quantidade.get()
    categoria = entrada_categoria.get()

    # Entrada na linha 2 e aumentando progressivamente
    proxima_linha = 2
    while vendas.cell(row=proxima_linha, column=1).value:  # Verifica se a célula está vazia
        proxima_linha += 1

    # Adiciona os dados à próxima linha vazia da planilha
    vendas.cell(row=proxima_linha, column=1, value=cliente)
    vendas.cell(row=proxima_linha, column=2, value=produto)
    vendas.cell(row=proxima_linha, column=3, value=quantidade)
    vendas.cell(row=proxima_linha, column=4, value=categoria)

    # Salva as alterações no arquivo Excel
    workbook.save('Vazio.xlsx')

    # Limpa os campos de entrada após salvar os dados
    entrada_cliente.delete(0, 'end')
    entrada_produto.delete(0, 'end')
    entrada_quantidade.delete(0, 'end')
    entrada_categoria.delete(0, 'end')

    messagebox.showinfo('Sucesso!', 'Cadastro Salvo com Sucesso')


root = tk.Tk()
root.geometry("225x150")
root.title("Cadastrar")

root.grid_columnconfigure(0, minsize=150, weight=1)  # Tamanho mínimo para a coluna 0

#Texto
label_cliente = tk.Label(root, text="Cliente:")
label_cliente.grid(row=0, column=0, sticky=tk.W)  # Alinha à esquerda

#Inserir
entrada_cliente = tk.Entry(root)
entrada_cliente.grid(row=0, column=1, sticky=tk.W)

#Texto
label_produto = tk.Label(root, text="Produto:")
label_produto.grid(row=1, column=0, sticky=tk.W)

#Inserir
entrada_produto = tk.Entry(root)
entrada_produto.grid(row=1, column=1, sticky=tk.W)

#Texto
label_quantidade = tk.Label(root, text="Quantidade:")
label_quantidade.grid(row=2, column=0, sticky=tk.W)

#Inserir
entrada_quantidade = tk.Entry(root)
entrada_quantidade.grid(row=2, column=1, sticky=tk.W)

#Texto
label_categoria = tk.Label(root, text="Categoria do Produto:")
label_categoria.grid(row=3, column=0, sticky=tk.W)

#Inserir
entrada_categoria = tk.Entry(root)
entrada_categoria.grid(row=3, column=1, sticky=tk.W)

botao_salvar = tk.Button(root, text="Salvar", command=salvar_dados, bg='#6495ED',fg="white")
botao_salvar.grid(row=4, column=0, columnspan=1, pady=20, sticky=tk.EW)

root.mainloop()